"""Rewrite the harvested WDPA extract so its country column matches the rest.

WDPA labels a protected area with an ISO3 code; the settlement layer and the
country facts use full names. Left alone, every protected area promoted into a
site candidate lands in its own country bucket -- 'MAR' scoring separately from
'Morocco', with its own quota and its own normalisation maximum.

Also merges the harvest parts, since the browser-side harvest was saved in
batches, and drops duplicates by WDPAID (transboundary areas appear in the file
of every country they touch).
"""
import csv
import os
from pathlib import Path

DATA = Path(os.environ.get("TWM_DATA", "/home/claude/twm/data"))


def main():
    import openpyxl
    wb = openpyxl.load_workbook(DATA / "Travelers World Map - Database Build.xlsx")
    ws = wb["Countries"]
    hdr = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
    ix = {h: i for i, h in enumerate(hdr)}
    iso3 = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, i3 = row[ix["country"]], row[ix["iso3"]]
        if name and i3:
            iso3[str(i3).strip().upper()] = str(name).strip()

    parts = sorted(DATA.glob("wdpa*.csv"))
    parts = [p for p in parts if p.name != "wdpa_normalised.csv"]
    seen, kept, dropped, unmapped = {}, 0, 0, {}
    for p in parts:
        with open(p, encoding="utf-8-sig", newline="") as fh:
            for row in csv.DictReader(fh):
                wid = row.get("WDPAID")
                if not wid or wid in seen:
                    continue
                code = (row.get("ISO3") or "").strip().upper()
                name = iso3.get(code)
                if not name:
                    unmapped[code] = unmapped.get(code, 0) + 1
                    dropped += 1
                    continue
                row["ISO3"] = name
                seen[wid] = row
                kept += 1

    out = DATA / "wdpa_normalised.csv"
    with open(out, "w", encoding="utf-8", newline="") as fo:
        w = csv.DictWriter(fo, fieldnames=["WDPAID", "NAME", "IUCN_CAT", "GIS_AREA",
                                           "ISO3", "latitude", "longitude"])
        w.writeheader()
        for row in seen.values():
            w.writerow({k: row.get(k, "") for k in w.fieldnames})

    print(f"wdpa: merged {len(parts)} part(s) -> {kept} unique areas, "
          f"{dropped} dropped for an unmapped ISO3")
    if unmapped:
        print("  unmapped:", sorted(unmapped.items(), key=lambda kv: -kv[1])[:12])
    countries = {r["ISO3"] for r in seen.values()}
    print(f"  covering {len(countries)} countries")


if __name__ == "__main__":
    main()
