"""Normalise raw sources into the shapes twm's adapters expect.

Nothing here invents data. It renames columns, repairs XML entity escaping, and
expands transnational World Heritage inscriptions to one point per state party.
Every output row is traceable to an input row.
"""
import csv
import html.entities
import json
import os
import re
import xml.etree.ElementTree as ET
from pathlib import Path

DATA = Path(os.environ.get("TWM_DATA", "/home/claude/twm/data"))
DATA.mkdir(parents=True, exist_ok=True)

ENT = re.compile(r"&([a-zA-Z][a-zA-Z0-9]{1,30});")
KEEP = {"amp", "lt", "gt", "quot", "apos"}


def _ent(m):
    n = m.group(1)
    if n in KEEP:
        return m.group(0)
    ch = html.entities.html5.get(n + ";")
    return "" if ch is None else ch.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# short_description and justification carry unbalanced inline HTML.
# The adapter never reads either field.
DESC = re.compile(r"<short_description>.*?</short_description>", re.DOTALL)
JUST = re.compile(r"<justification>.*?</justification>", re.DOTALL)


def clean_xml(path: Path) -> ET.Element:
    raw = path.read_text(encoding="utf-8", errors="replace")
    raw = DESC.sub("<short_description/>", raw)
    raw = JUST.sub("<justification/>", raw)
    return ET.fromstring(ENT.sub(_ent, raw))


def main():
    # ------------------------------------------------------ country lookups
    import openpyxl
    wb = openpyxl.load_workbook(DATA / "Travelers World Map - Database Build.xlsx")
    ws = wb["Countries"]
    hdr = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
    ix = {h: i for i, h in enumerate(hdr)}

    iso2_to_country, countries = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, iso2, area = row[ix["country"]], row[ix["iso"]], row[ix["area km2"]]
        if not name:
            continue
        name = str(name).strip()
        if iso2:
            iso2_to_country[str(iso2).strip().lower()] = name
        try:
            area = float(area)
        except (TypeError, ValueError):
            continue
        if area > 0:
            countries[name] = {"area_km2": area, "iso": str(iso2 or "").strip()}
    (DATA / "countries.json").write_text(
        json.dumps(countries, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"[countries]   {len(countries)} with an area")

    # -------------------------------------------------------------- 1. WHS
    root = clean_xml(DATA / "whs_current.xml")
    rows = root.findall(".//row")
    out = ET.Element("query")
    n_expanded = n_sites = n_nogeo = 0
    def text_of(elem, tag):
        e = elem.find(tag)
        return e.text.strip() if e is not None and e.text else ""

    for r in rows:
        def t(tag, _r=r):
            return text_of(_r, tag)

        uid = t("id_number") or t("unique_number")
        pts = {}
        for poi in r.findall("./geolocations/poi"):
            la, lo = poi.findtext("latitude"), poi.findtext("longitude")
            iso = (poi.findtext("iso2") or "").strip().lower()
            if not la or not lo or iso in pts:
                continue
            pts[iso] = (la, lo)
        if not pts:
            la, lo = t("latitude"), t("longitude")
            if not la or not lo:
                n_nogeo += 1
                continue
            pts[t("iso_code").split(",")[0].strip().lower()] = (la, lo)
        n_sites += 1
        for iso, (la, lo) in pts.items():
            e = ET.SubElement(out, "row")
            for tag, val in (("id_number", uid), ("unique_number", t("unique_number")),
                             ("site", t("site") or t("name_en")), ("category", t("category")),
                             ("danger", t("danger")), ("latitude", la), ("longitude", lo),
                             ("states", iso2_to_country.get(iso, t("states")))):
                ET.SubElement(e, tag).text = val
            n_expanded += 1
    out.set("rows", str(n_expanded))
    ET.ElementTree(out).write(DATA / "whs_expanded.xml", encoding="utf-8", xml_declaration=True)
    print(f"[whs]         {len(rows)} inscriptions -> {n_sites} with coordinates "
          f"-> {n_expanded} rows (one per state party); {n_nogeo} dropped for no coordinates")

    # ------------------------------------------------- 2. settlement layer
    n = 0
    with open(DATA / "world_settlements.csv", encoding="utf-8-sig", newline="") as fh, \
         open(DATA / "settlements_ucdb.csv", "w", encoding="utf-8", newline="") as fo:
        w = csv.DictWriter(fo, fieldnames=["ID_UC_G0", "name", "country", "population",
                                           "latitude", "longitude"])
        w.writeheader()
        for row in csv.DictReader(fh):
            try:
                lat, lon = float(row["lat"]), float(row["lon"])
                pop = float(row["population"] or 0)
            except (TypeError, ValueError, KeyError):
                continue
            w.writerow({"ID_UC_G0": row["geonameid"], "name": row["settlement"],
                        "country": row["country"], "population": pop,
                        "latitude": lat, "longitude": lon})
            n += 1
    print(f"[settlements] {n} rows")


if __name__ == "__main__":
    main()
