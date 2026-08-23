"""Build the printed map's territory tiles.

Territories are the physical objects: magnetic tiles cut from merged
administrative boundaries, each carrying a workable number of places. This runs
`twm.territories` unchanged -- the only local part is loading Natural Earth's
admin-1 layer from a GeoJSON mirror, because the upstream CDN is unreachable
from this machine.

Country names are matched by ISO 3166-1 alpha-2, not by label: Natural Earth
says "United States of America" where the place database says "United States",
and matching on text would silently drop a country's entire tile set.
"""
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, os.environ.get("TWM_PKG", "/home/claude/twm/db"))

from twm.geo import PrintedMap
from twm.territories import (
    AdminUnit,
    assign_places,
    build_territories,
    dissolve_disputed,
)
from twm.types import Place

DATA = Path(os.environ.get("TWM_DATA", "/home/claude/twm/data"))
DIST = Path(os.environ.get("TWM_DIST", "/home/claude/twm/dist"))


def load_admin_units(iso2_to_country):
    """Natural Earth admin-1 -> AdminUnit, keyed to our country names by ISO2."""
    from shapely.geometry import shape

    raw = json.loads((DATA / "ne_10m_admin_1.geojson").read_text(encoding="utf-8"))
    units, skipped = [], 0
    for f in raw["features"]:
        p = f["properties"]
        iso2 = (p.get("iso_a2") or "").strip().upper()
        country = iso2_to_country.get(iso2.lower())
        if not country:
            skipped += 1
            continue
        geom = f.get("geometry")
        if not geom:
            skipped += 1
            continue
        try:
            g = shape(geom)
        except Exception:                                          # noqa: BLE001
            skipped += 1
            continue
        if g.is_empty:
            skipped += 1
            continue
        if not g.is_valid:
            g = g.buffer(0)
        units.append(AdminUnit(
            unit_id=p.get("adm1_code") or f"{iso2}-{len(units)}",
            name=(p.get("name_en") or p.get("name") or "").strip() or "unnamed",
            country=country,
            geometry=g,
            level=1,
        ))
    print(f"admin units      {len(units):>7}   ({skipped} skipped: no ISO match or no geometry)")
    return units


def _iso3_by_country():
    import openpyxl
    wb = openpyxl.load_workbook(DATA / "Travelers World Map - Database Build.xlsx")
    ws = wb["Countries"]
    hdr = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows())]
    ix = {h: i for i, h in enumerate(hdr)}
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, i3 = row[ix["country"]], row[ix["iso3"]]
        if name and i3:
            out[str(name).strip()] = str(i3).strip().upper()
    return out


def _reid_on_iso3(terrs):
    iso3 = _iso3_by_country()
    seen, fallback = {}, {}
    for t in terrs:
        code = iso3.get(t.country)
        if not code:
            # no ISO3 on record: derive a stable 3-char code, disambiguated
            base = "".join(ch for ch in t.country.upper() if ch.isalnum())[:3] or "XXX"
            n = fallback.setdefault(base, 0) + 1
            fallback[base] = n
            code = base if n == 1 else f"{base[:2]}{n}"
        i = seen.get(code, 0) + 1
        seen[code] = i
        t.territory_id = f"{code}-T{i:02d}"
    ids = [t.territory_id for t in terrs]
    assert len(ids) == len(set(ids)), "territory ids still not unique"
    return terrs


def main():
    countries_raw = json.loads((DATA / "countries.json").read_text(encoding="utf-8"))
    iso2 = {v.get("iso", "").lower(): k for k, v in countries_raw.items() if v.get("iso")}

    app = json.loads((DIST / "app_places.json").read_text(encoding="utf-8"))
    places = [Place(
        place_id=p["place_id"], name=p["name"], country=p["country"],
        lat=p["lat"], lon=p["lon"], is_site=p["is_site"], score=p["score"],
        archetypes=p["archetypes"], archetype_weights=p["archetype_weights"],
        whs=p.get("whs", 0), reach=p.get("reach", "near"),
        best_months=p.get("best_months", []),
        on_printed_map=p["on_printed_map"], printed_rank=p.get("printed_rank"),
        sources=p.get("sources", []),
    ) for p in app["places"]]
    printed = [p for p in places if p.on_printed_map]
    print(f"places           {len(places):>7}   on the printed map {len(printed)}")

    units = dissolve_disputed(load_admin_units(iso2))
    units = assign_places(units, printed)
    with_places = sum(1 for u in units if u.place_ids)
    print(f"units with places{with_places:>7}")

    pm = PrintedMap()
    terrs = build_territories(units, printed, printed_map=pm)

    # `territories._slug` takes the first three characters of the country name,
    # which is not unique: Australia and Austria both slug to AUS, India and
    # Indonesia to IND, and all four Saint-somethings to SAI. DuckDB's primary
    # key catches it. Territory ids are a stability contract like place ids, so
    # renumber them on ISO 3166-1 alpha-3, which is unique by construction.
    terrs = _reid_on_iso3(terrs)

    printable = [t for t in terrs if t.printable]
    print(f"territories      {len(terrs):>7}   printable {len(printable)} "
          f"(min tile extent {pm.min_tile_extent_km:.0f} km)")

    payload = {
        "territories": [{
            "territory_id": t.territory_id, "name": t.name, "country": t.country,
            "place_ids": t.place_ids, "admin_units": t.admin_units,
            "dominant_archetypes": t.dominant_archetypes, "printable": t.printable,
            "places": len(t.place_ids),
        } for t in terrs],
        "min_tile_extent_km": round(pm.min_tile_extent_km, 1),
        "map_width_m": pm.width_m,
    }
    (DIST / "territories.json").write_text(
        json.dumps(payload, separators=(",", ":")), encoding="utf-8")

    try:
        from twm.store import Store
        st = Store(DIST / "twm.duckdb")
        st.write_territories(terrs)
        st.close()
        print("territories written to twm.duckdb")
    except Exception as exc:                                       # noqa: BLE001
        print(f"duckdb territories skipped: {exc}")

    by_country = {}
    for t in terrs:
        by_country.setdefault(t.country, []).append(t)
    top = sorted(by_country.items(), key=lambda kv: -len(kv[1]))[:8]
    print("\nmost tiles:")
    for c, ts in top:
        print(f"  {c:<22} {len(ts):>3} tiles, "
              f"{sum(len(t.place_ids) for t in ts):>3} places, "
              f"{sum(1 for t in ts if t.printable):>3} printable")
    return terrs


if __name__ == "__main__":
    main()
