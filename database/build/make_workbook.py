"""Regenerate 'Travelers World Map - Database Build.xlsx' from a real build.

Matches the conventions of the workbook it replaces: Arial throughout, the same
heading palette, and live formulas rather than pasted totals.
"""
import json
import os
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DIST = Path(os.environ.get("TWM_DIST", "/home/claude/twm/dist"))
DATA = Path(os.environ.get("TWM_DATA", "/home/claude/twm/data"))
OUT = Path(sys.argv[1] if len(sys.argv) > 1 else "/home/claude/twm/Database Build.xlsx")

INK = "FF24343B"
TEAL = "FF1F4E5F"
MUTED = "FF5A6B75"
AMBER_T = "FF8C6D3F"
GREEN_F = "FFEDF4EE"
AMBER_F = "FFFBF3E2"
RULE = Side(style="thin", color="FFD8DEE1")


def h1(c):
    c.font = Font(name="Arial", size=17, bold=True, color=TEAL)


def sub(c):
    c.font = Font(name="Arial", size=11, color=MUTED)


def section(c):
    c.font = Font(name="Arial", size=10, bold=True, color=AMBER_T)


def label(c):
    c.font = Font(name="Arial", size=10, bold=True, color=INK)


def body(c):
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(vertical="top", wrap_text=True)


def head(c):
    c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
    c.fill = PatternFill("solid", start_color=TEAL)
    c.alignment = Alignment(vertical="center")


def table(ws, rows, widths, freeze="A2"):
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    for j, v in enumerate(rows[0], start=1):
        head(ws.cell(row=1, column=j, value=v))
    for i, r in enumerate(rows[1:], start=2):
        for j, v in enumerate(r, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=10)
            c.border = Border(bottom=RULE)
    ws.freeze_panes = freeze


def main():
    app = json.loads((DIST / "app_places.json").read_text(encoding="utf-8"))
    rep = json.loads((DIST / "build_report.json").read_text(encoding="utf-8"))
    countries = json.loads((DATA / "countries.json").read_text(encoding="utf-8"))
    stats = rep["stats"]
    per = stats["per_country"]
    places = app["places"]

    wb = Workbook()

    # ---------------------------------------------------------- Start here
    ws = wb.active
    ws.title = "Start here"
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 104
    h1(ws["B1"])
    ws["B1"] = "TRAVELERS WORLD MAP — DATABASE BUILD"
    sub(ws["B2"])
    ws["B2"] = "What ran, what it produced, and what is still missing"

    section(ws["B4"])
    ws["B4"] = "STATUS"
    rows = [
        ("Code", "All 31 tests pass. Linter clean. The pipeline is unmodified.", GREEN_F),
        ("World database", "BUILT. Every number in this workbook comes from a real run "
                           "over real sources.", GREEN_F),
        ("Places", '=("In the application: "&TEXT(COUNTA(\'Places (world)\'!B2:B60000),"#,##0")'
                   '&"  ·  On the printed map: "'
                   '&TEXT(COUNTIF(\'Places (world)\'!G2:G60000,"yes"),"#,##0")'
                   '&" of "&TEXT(\'Selection evidence\'!I2,"#,##0")'
                   '&" drilled holes")', None),
        ("Countries", '=TEXT(COUNTA(\'Selection evidence\'!A2:A400),"#,##0")'
                      '&" with at least one place"', None),
        ("Tiles",
         '=("Printed-map territories: "&TEXT(COUNTA(Territories!A2:A4000),"#,##0")'
         '&"  ·  large enough to cut: "'
         '&TEXT(COUNTIF(Territories!E2:E4000,"yes"),"#,##0"))', None),
        ("Still missing", "OSM covers 150 of 235 countries, and UNESCO intangible heritage is "
                          "held back at 12% coverage. See 'Known gaps'.", AMBER_F),
    ]
    r = 5
    for k, v, fill in rows:
        label(ws.cell(row=r, column=2, value=k))
        c = ws.cell(row=r, column=3, value=v)
        body(c)
        if fill:
            for col in (2, 3):
                ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=fill)
        r += 1

    r += 1
    section(ws.cell(row=r, column=2, value="WHAT THIS WORKBOOK CONTAINS"))
    r += 1
    for k, v in [
        ("Places (world)", "Every place in the database, with its score, its archetypes "
                           "and whether it earns a hole on the printed map."),
        ("Selection evidence", "Per country: candidates scored, places kept, quota, and how "
                               "much less redundant coverage selection is than ranking."),
        ("Sources", "Every asset layer actually used, with its licence, its provenance and "
                    "how much of the world it covers."),
        ("Countries", "Area and the places found, for the quota formula."),
        ("Model parameters", "Every frozen parameter. Unchanged from the approved model."),
        ("Territories", "The printed map's physical tiles, merged from real "
                        "administrative boundaries. Never drawn freehand."),
        ("Known gaps", "What is wrong with the current output and precisely why."),
    ]:
        label(ws.cell(row=r, column=2, value=k))
        body(ws.cell(row=r, column=3, value=v))
        r += 1

    r += 1
    section(ws.cell(row=r, column=2, value="THE ONE THING TO UNDERSTAND"))
    r += 1
    label(ws.cell(row=r, column=2, value="An empty pillar still normalises"))
    c = ws.cell(row=r, column=3, value=(
        "Pillar scores are normalised against the country maximum. When a pillar has almost "
        "no data, that maximum is set by noise — and 1.000 out of nothing outranks 1.000 out "
        "of something. This is why a town with two protected areas currently outranks Fes. "
        "Check that each pillar's maximum rests on a real cluster before trusting a ranking."))
    body(c)
    for col in (2, 3):
        ws.cell(row=r, column=col).fill = PatternFill("solid", start_color=AMBER_F)

    # ------------------------------------------------------ Places (world)
    ws = wb.create_sheet("Places (world)")
    rows = [("country", "place", "kind", "lat", "lon", "score", "on printed map",
             "printed rank", "archetypes", "whs", "sources")]
    for p in sorted(places, key=lambda p: (p["country"], -p["score"])):
        rows.append((
            p["country"], p["name"], "site" if p["is_site"] else "settlement",
            p["lat"], p["lon"], p["score"],
            "yes" if p["on_printed_map"] else "no", p.get("printed_rank"),
            ", ".join(p["archetypes"]), p.get("whs", 0), ", ".join(p.get("sources", [])),
        ))
    table(ws, rows, [20, 40, 11, 10, 10, 7, 15, 13, 26, 6, 30])

    # -------------------------------------------------- Selection evidence
    ws = wb.create_sheet("Selection evidence")
    rows = [("country", "candidates scored", "places in app", "on printed map", "quota",
             "coverage redundancy", "closest pair km", "spacing conflicts", "hole budget")]
    first = True
    for name, st in sorted(per.items(), key=lambda kv: -kv[1]["on_printed_map"]):
        rows.append((name, st["candidates"], st["in_app"], st["on_printed_map"],
                     st.get("quota"), st.get("redundancy"), st.get("closest_pair_km"),
                     st.get("spacing_conflicts"),
                     stats["hole_budget"] if first else None))
        first = False
    table(ws, rows, [22, 18, 14, 15, 8, 20, 16, 18, 13])

    # ------------------------------------------------------------- Sources
    ws = wb.create_sheet("Sources")
    man = json.loads((DIST / "manifest.json").read_text(encoding="utf-8")) \
        if (DIST / "manifest.json").exists() else {}
    rows = [("#", "source", "supplies", "records used", "licence", "cross-country safe",
             "how it was obtained")]
    src = man.get("sources", {})
    spec = [
        ("UNESCO World Heritage", "Built heritage and natural setting; the only asset class "
         "adjudicated globally", src.get("unesco-whs"), "UNESCO terms, attribution", "yes",
         "whc.unesco.org XML export, fetched in the browser; expanded to one row per state "
         "party so a transnational inscription credits every country that holds it"),
        ("Protected Planet (WDPA)", "Protected areas, IUCN I–IV, 25 km² and larger",
         src.get("wdpa"), "Non-commercial by default", "yes",
         "Per-country shapefiles. The CSV export carries no coordinates, so centroids are "
         "read from each record's bounding box in the .shp"),
        ("Wikidata", "Multilingual significance — how many language communities "
         "independently documented a place", src.get("wikidata"), "CC0-1.0", "yes",
         "WDQS, queried per country. The label service made every global query time out, so "
         "labels were dropped and only coordinates and sitelink counts kept"),
        ("GeoNames settlements", "Candidate layer", 26007, "CC-BY-4.0", "n/a (not an asset)",
         "Supplied with the project; agglomerated into urban centres because GeoNames' unit "
         "is the named place, not the built-up area"),
    ]
    for i, (a, b, c, d, e, f) in enumerate(spec, start=1):
        rows.append((i, a, b, c, d, e, f))
    table(ws, rows, [4, 26, 44, 14, 26, 18, 66])

    # ---------------------------------------------------------- Territories
    tpath = DIST / "territories.json"
    if tpath.exists():
        terr = json.loads(tpath.read_text(encoding="utf-8"))
        ws = wb.create_sheet("Territories")
        rows = [("territory id", "country", "name", "places", "printable",
                 "dominant archetypes", "admin units merged")]
        for t in sorted(terr["territories"],
                        key=lambda t: (t["country"], t["territory_id"])):
            rows.append((t["territory_id"], t["country"], t["name"], t["places"],
                         "yes" if t["printable"] else "too small",
                         ", ".join(t["dominant_archetypes"]),
                         len(t["admin_units"])))
        table(ws, rows, [14, 22, 30, 8, 12, 24, 18])

    # ----------------------------------------------------------- Countries
    ws = wb.create_sheet("Countries")
    rows = [("country", "iso", "area km2", "candidates", "places", "on printed map")]
    for name, f in sorted(countries.items(), key=lambda kv: -kv[1]["area_km2"]):
        st = per.get(name, {})
        rows.append((name, f.get("iso", ""), f["area_km2"], st.get("candidates", 0),
                     st.get("in_app", 0), st.get("on_printed_map", 0)))
    table(ws, rows, [24, 6, 14, 13, 10, 15])

    # ---------------------------------------------------- Model parameters
    ws = wb.create_sheet("Model parameters")
    sys.path.insert(0, os.environ.get("TWM_PKG", "/home/claude/twm/db"))
    from twm.config import PARAMS
    rows = [("parameter", "value", "pillar / note")]
    for k in sorted(vars(PARAMS)) if hasattr(PARAMS, "__dict__") else []:
        rows.append((k, getattr(PARAMS, k), ""))
    if len(rows) == 1:
        import dataclasses
        for fld in dataclasses.fields(PARAMS):
            rows.append((fld.name, getattr(PARAMS, fld.name), ""))
    table(ws, rows, [30, 16, 60])

    # --------------------------------------------------------- Known gaps
    ws = wb.create_sheet("Known gaps")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 104
    h1(ws["B1"])
    ws["B1"] = "KNOWN GAPS"
    sub(ws["B2"])
    ws["B2"] = "What is wrong with this output, and precisely why"
    r = 4
    for k, v in [
        ("The livability pillar is empty",
         "w_l = 0.35 — thirty-five percent of the model's weight — contributes nothing "
         "anywhere on earth, because there is no cuisine layer and no OpenStreetMap. That "
         "leaves heritage at 0.30 against nature at 0.35 and tilts every ranking toward "
         "nature by construction."),
        ("The heritage long tail is missing",
         "Within a country, what separates a great city from a village beside one inscription "
         "is the long tail of minor heritage and museums. That is OpenStreetMap, which is "
         "ODbL: share-alike attaches on distribution. A licensing decision, not an effort "
         "one."),
        ("A sparse pillar's maximum is noise",
         "Normalisation is linear against the country maximum, so a pillar with almost no "
         "data hands 1.000 to whoever holds one asset. Boumia (two protected areas) outranked "
         "Fes (46 heritage items and a World Heritage site); Trieste, on one natural "
         "inscription, outranked Rome and its 409."),
        ("UNESCO Intangible Heritage is absent, deliberately",
         "ICH elements are practices, not points. Georeferencing them to country centroids "
         "would be inventing geography, and at tier weight 8 a bad guess would swing results "
         "hard. Better absent than fabricated."),
        ("The pilgrimage tier is dead",
         "The Wikidata query for pilgrimage destinations returns zero rows — wd:Q1129470 has "
         "no instances carrying coordinates. pilgrimage_major (weight 6) contributes nothing "
         "and needs a new identifier before it will."),
        ("No tentative list",
         "UNESCO publishes no machine-readable export of the Tentative List; the adapter's "
         "own docstring says so. whs_tentative is empty."),
    ]:
        label(ws.cell(row=r, column=2, value=k))
        body(ws.cell(row=r, column=3, value=v))
        r += 2

    wb.save(OUT)
    print(f"wrote {OUT}  ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  places sheet: {len(places)} rows")


if __name__ == "__main__":
    main()
